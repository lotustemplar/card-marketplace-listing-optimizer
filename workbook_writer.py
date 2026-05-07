from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from pricing_logic import DISPLAY_COLUMNS_DIRECT, DISPLAY_COLUMNS_MANAPOOL, ProcessResult


HEADER_FILL = PatternFill(fill_type="solid", fgColor="0F4C5C")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FORCED_MIN_FILL = PatternFill(fill_type="solid", fgColor="FFF3BF")
ALERT_FILL = PatternFill(fill_type="solid", fgColor="F8D7DA")
CLIFF_FILL = PatternFill(fill_type="solid", fgColor="FFE8CC")

CURRENCY_COLUMNS = {
    "Manapool Sheet": {"Manapool Price", "Manapool Net", "Base Direct Price", "Base Direct Net", "Required Direct Price"},
    "TCGPlayer Direct Sheet": {"Direct Listing Price", "Direct Net", "Manapool Price", "Manapool Net"},
    "Analysis": {"Value"},
}
PERCENT_COLUMNS = {
    "Manapool Sheet": {"Direct Bump %"},
    "TCGPlayer Direct Sheet": {"Direct Bump %"},
}


def write_dataframe(ws, dataframe: pd.DataFrame) -> None:
    if dataframe.empty:
        for col_idx, column_name in enumerate(dataframe.columns, start=1):
            ws.cell(row=1, column=col_idx, value=column_name)
        return

    for col_idx, column_name in enumerate(dataframe.columns, start=1):
        ws.cell(row=1, column=col_idx, value=column_name)

    for row_idx, row in enumerate(dataframe.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def format_header(ws) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    if ws.max_column and ws.max_row:
        ws.auto_filter.ref = ws.dimensions


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def apply_sheet_formats(ws, sheet_name: str) -> None:
    headers = [cell.value for cell in ws[1]]
    header_index = {header: idx + 1 for idx, header in enumerate(headers)}

    for row_idx in range(2, ws.max_row + 1):
        for column_name in CURRENCY_COLUMNS.get(sheet_name, set()):
            col_idx = header_index.get(column_name)
            if col_idx and isinstance(ws.cell(row=row_idx, column=col_idx).value, (int, float)):
                ws.cell(row=row_idx, column=col_idx).number_format = "$#,##0.00"

        for column_name in PERCENT_COLUMNS.get(sheet_name, set()):
            col_idx = header_index.get(column_name)
            if col_idx and isinstance(ws.cell(row=row_idx, column=col_idx).value, (int, float)):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.0%"

    if sheet_name == "Analysis":
        metric_col = header_index.get("Metric")
        value_col = header_index.get("Value")
        for row_idx in range(2, ws.max_row + 1):
            metric_value = ws.cell(row=row_idx, column=metric_col).value if metric_col else None
            value_cell = ws.cell(row=row_idx, column=value_col) if value_col else None
            if not value_cell or metric_value is None:
                continue
            metric_text = str(metric_value)
            if any(token in metric_text for token in ["Total estimated", "Combined estimated", "Shipping/Supply Cost", "Threshold", "Cost"]):
                if isinstance(value_cell.value, (int, float)):
                    value_cell.number_format = "$#,##0.00"
            elif "Average Direct bump %" in metric_text and isinstance(value_cell.value, (int, float)):
                value_cell.number_format = "0.0%"


def highlight_special_rows(ws, dataframe: pd.DataFrame) -> None:
    headers = [cell.value for cell in ws[1]]
    header_index = {header: idx + 1 for idx, header in enumerate(headers)}

    if ws.title == "Manapool Sheet":
        price_col = header_index.get("Manapool Price")
        bump_col = header_index.get("Direct Bump %")
        for row_idx, row in enumerate(dataframe.to_dict("records"), start=2):
            if row.get("_cliff_affected"):
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = CLIFF_FILL
            if row.get("_forced_min") and price_col:
                ws.cell(row=row_idx, column=price_col).fill = FORCED_MIN_FILL
            if row.get("_bump_exceeded") and bump_col:
                ws.cell(row=row_idx, column=bump_col).fill = ALERT_FILL

    if ws.title == "TCGPlayer Direct Sheet":
        bump_col = header_index.get("Direct Bump %")
        for row_idx, row in enumerate(dataframe.to_dict("records"), start=2):
            if row.get("_cliff_affected"):
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = CLIFF_FILL
            if row.get("_bump_exceeded") and bump_col:
                ws.cell(row=row_idx, column=bump_col).fill = ALERT_FILL


def build_workbook(result: ProcessResult) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    manapool_sheet = workbook.create_sheet("Manapool Sheet")
    direct_sheet = workbook.create_sheet("TCGPlayer Direct Sheet")
    analysis_sheet = workbook.create_sheet("Analysis")
    errors_sheet = workbook.create_sheet("Errors")

    manapool_output = result.manapool_full_df[DISPLAY_COLUMNS_MANAPOOL] if not result.manapool_full_df.empty else pd.DataFrame(columns=DISPLAY_COLUMNS_MANAPOOL)
    direct_output = result.direct_full_df[DISPLAY_COLUMNS_DIRECT] if not result.direct_full_df.empty else pd.DataFrame(columns=DISPLAY_COLUMNS_DIRECT)

    write_dataframe(manapool_sheet, manapool_output)
    write_dataframe(direct_sheet, direct_output)
    write_dataframe(analysis_sheet, result.analysis_df)
    write_dataframe(errors_sheet, result.errors_df if not result.errors_df.empty else pd.DataFrame(columns=["Error reason"]))

    for sheet in [manapool_sheet, direct_sheet, analysis_sheet, errors_sheet]:
        format_header(sheet)
        apply_sheet_formats(sheet, sheet.title)
        autosize_columns(sheet)

    highlight_special_rows(manapool_sheet, result.manapool_full_df)
    highlight_special_rows(direct_sheet, result.direct_full_df)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
